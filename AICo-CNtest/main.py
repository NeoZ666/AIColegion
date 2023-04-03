from email.message import EmailMessage
import ssl
import smtplib
import openpyxl
import os
from dotenv import load_dotenv

load_dotenv()

wb = openpyxl.load_workbook(r"E:\Firefox Downloads\test_ranks_2023-03-04T13_11_24.312136+05_30(1).xlsx")

sh = wb.active

email_sender = "ai_ds.aicolegion@ves.ac.in"
email_password = os.getenv("mail_passcode")
em = EmailMessage()
em["From"] ="AI CoLegion - VESIT <ai_ds.colegion@ves.ac.in>"
subj = "Results of Coding Ninja: CoLegion Coding Test"

em["Subject"] = subj
context = ssl.create_default_context()
# email_receiver = '2020.naresh.shewkani@ves.ac.in'
for i in range(3,59):    #first index is inclusive, last index is exclusive
        def html_text():
            em.add_alternative(
                f"""\
                <!DOCTYPE html>
                <html>
                    <body>
                            <p>Dear {name}!</p>
                            <br>
                            We hope this email finds you well!
                            <br>
                            We would like to take this opportunity to thank you for taking out your time and participate in the <b>Coding Ninja: CoLegion Coding Contest</b> in association with <b>AI CoLegion - VESIT</b>. We had an overwhelming response to the contest with over <b>350+</b> participants and we are delighted to see so many talented coders competing.
                            <br>
                            We are pleased to announce the results of the contest. Congratulations to all the winners! Here are the details of your performance:
                            <br>
                            <br>
                            <b>Score:</b> {Score}
                            <br>
                            <b>Rank:</b> {Rank}
                            <br>
                            <b>Duration:</b> {Duration} minutes
                            <br>
                            <br>
                            We appreciate your hard work and dedication, and we hope that this achievement will inspire you to keep pursuing your passion for coding. We are grateful for your contribution.
                            Thank you for being a part of our event. We hope you enjoyed the experience and learned something new.
                            <br>
                            <br>
                            For any further queries reply back to this mail or contact☎️:
                            <br>
                            <br>
                            Naresh Shewkani: 8483030123
                            <br>
                            Priyanshu Singh: 9082035567
                            <br>
                            ---
                            <br>
                            Regards,
                            <br>
                            <b>Team AI CoLegion - VESIT</b>
                            <br>
                            <br>
                            Follow our Instagram and LinkedIn pages for further updates regarding future events:
                            <br> 
                            <a href="https://www.instagram.com/aicolegion_vesit/"> <img src="https://i.imgur.com/l6kR6Av.png" alt ="Instagram" style="width: 40px; height:40px;" /> </a>
                            &nbsp;
                            <a href="https://www.linkedin.com/company/ai-colegion-vesit/">
                                            <img src="https://upload.wikimedia.org/wikipedia/commons/thumb/8/81/LinkedIn_icon.svg/120px-LinkedIn_icon.svg.png" alt="LinkedIn" style="width:40px; height:40px;" />
                            </a>

                        </p>
                    </body>
                </html>
            """,
                subtype="html",
            )
        email_receiver = sh.cell(row=i,column=2).value
        name = sh.cell(row=i,column=3).value.capitalize()
        Score = sh.cell(row=i,column=4).value
        Duration = sh.cell(row=i, column=8).value
        Rank = sh.cell(row=i,column=1).value
        html_text()
        # pdf_path = f'generated_pdf/{names}.pdf'
        # with open(pdf_path, 'rb') as f:
        #     file_data = f.read()
        #     file_name = names + '.pdf'
        # em.add_attachment(file_data, maintype='application',subtype='octet-stream',filename=file_name)
        em['To'] = email_receiver
        with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
            # with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
            #     smtp.starttls()
            smtp.login(email_sender, email_password)
            smtp.sendmail(email_sender, email_receiver, em.as_string())
            print(f'Message sent to {name} with the index {i}')
            em.clear_content()
        del em['To']