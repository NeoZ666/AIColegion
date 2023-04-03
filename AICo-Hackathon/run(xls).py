import cv2
import openpyxl
from fpdf import FPDF

template = 'certi_template.png'
issued_to = 'participants.xlsx'
output_path = 'generated_certi_data'

obj = openpyxl.load_workbook(issued_to)
sheet = obj.active

for i in range(1,215):
    if sheet.cell(row=i, column=2).value is None:
        pass
    else:
        named = sheet.cell(row=i, column=1)
        name = named.value
        img = cv2.imread(template)
        # if len(name) in range(8,15):
        #     cv2.putText(img, name, (410, 490), cv2.FONT_HERSHEY_TRIPLEX, 2, (255, 255, 0), 2, cv2.LINE_AA)
        # elif len(name) in range(14,19):
        #     cv2.putText(img, name, (335, 490), cv2.FONT_HERSHEY_TRIPLEX, 2, (255, 255, 0), 2, cv2.LINE_AA)
        # elif len(name) > 18:
        #     cv2.putText(img, name, (315, 490), cv2.FONT_HERSHEY_TRIPLEX, 1.65, (255, 255, 0), 2, cv2.LINE_AA)
        # else:
        #     cv2.putText(img, name, (445, 490), cv2.FONT_HERSHEY_TRIPLEX, 2, (255, 255, 0), 2, cv2.LINE_AA)
        middle_point = (989, 727) # only for reference 
        start_point = 989 - (len(name)*29)
        cv2.putText(img, name, (start_point, 743), cv2.FONT_HERSHEY_TRIPLEX, 3,(223,225,57), 3, cv2.LINE_AA)

        cv2.imwrite(f'generated_certi_data/{name}.png', img)
